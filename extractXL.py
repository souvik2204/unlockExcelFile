from openpyxl import load_workbook

def unprotect_sheets(file_path, output_path):
    try:
        # Load the workbook
        workbook = load_workbook(filename=file_path)
        print("Workbook loaded successfully.")
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return
    
    # Ensure there are sheets in the workbook
    if not workbook.sheetnames:
        print("No sheets found in the workbook.")
        return
    
    print(f"Sheet names: {workbook.sheetnames}")

    # Unprotect all sheets
    for ws in workbook.worksheets:
        print(f"Unprotecting sheet: {ws.title}")
        ws.protection.sheet = False
    
    try:
        # Save the workbook without sheet protection
        workbook.save(output_path)
        print(f"Sheet protection removed. File saved as {output_path}.")
    except Exception as e:
        print(f"Error saving workbook: {e}")

if __name__ == "__main__":
    file_path = "pythonScripts/xlfile.xlsx"  # Replace with your file path
    output_path = "pythonScripts/new/output.xlsx"  # Output file path
    
    unprotect_sheets(file_path, output_path)
